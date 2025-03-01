# core/reporter.py
# Professional Excel report generation with advanced styling

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from config.settings import REPORT_FILE
from utils.logger import logger
from utils.exceptions import ScannerError

def generate_report(vulnerabilities, results):
    wb = Workbook()
    ws = wb.active
    ws.title = "Vulnerability Verification Report"
    
    # Headers with advanced styling
    headers = ["IP", "Port", "Service", "Vulnerability", "Plugin ID", "Status", "PoC Folder Path"]
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A86E8", end_color="4A86E8", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    fills = {
        "Verified": PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),
        "False Positive": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        "Manual Check Required": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    }

    for vuln, (status, poc_path) in zip(vulnerabilities, results):
        row = [vuln["ip"], vuln["port"], vuln["service"], vuln["vulnerability"], vuln["plugin_id"], status, poc_path]
        ws.append(row)
        cell = ws.cell(row=ws.max_row, column=6)
        cell.fill = fills.get(status, fills["Manual Check Required"])
        for col_cell in ws[ws.max_row]:
            col_cell.border = thin_border

    # Auto-adjust column widths with padding
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    try:
        wb.save(REPORT_FILE)
        logger.info(f"\033[92mReport saved as {REPORT_FILE}\033[0m")
    except PermissionError:
        logger.error(f"\033[91mCannot save {REPORT_FILE}: File is open or permission denied\033[0m")
        raise ScannerError(f"Cannot save report: File is open or permission denied")
    except Exception as e:
        logger.error(f"\033[91mError saving report: {e}\033[0m")
        raise ScannerError(f"Error saving report: {e}")