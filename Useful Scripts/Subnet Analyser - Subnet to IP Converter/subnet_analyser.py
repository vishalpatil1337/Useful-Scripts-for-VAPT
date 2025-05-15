#!/usr/bin/env python3
"""
Subnet Analysis Tool
-------------------
A professional tool for analyzing IP subnet ranges and generating detailed reports.

This script reads subnet definitions from scope.txt and outputs the results to an Excel file
with enhanced formatting and detailed information.
"""

import ipaddress
import pandas as pd
import os
import sys
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def print_status(message):
    """Print status message to console."""
    print(f"[*] {message}")

def read_subnets(filename='scope.txt'):
    """Read subnets from file."""
    try:
        if not os.path.exists(filename):
            print(f"Error: Input file '{filename}' not found.")
            sys.exit(1)
            
        with open(filename, 'r') as file:
            subnets = [line.strip() for line in file if line.strip()]
            
        if not subnets:
            print("Error: No valid subnets found in the input file.")
            sys.exit(1)
            
        print_status(f"Found {len(subnets)} subnets in {filename}")
        return subnets
    except Exception as e:
        print(f"Error reading subnet file: {str(e)}")
        sys.exit(1)

def analyze_subnets(subnets):
    """Analyze all subnets and return results."""
    print_status("Analyzing subnets...")
    
    subnet_data = {}
    stats = {
        'total_subnets': len(subnets),
        'valid_subnets': 0,
        'invalid_subnets': 0,
        'total_ips': 0,
        'usable_ips': 0
    }
    
    subnet_summary = []
    
    for subnet in subnets:
        try:
            # Create network object
            network = ipaddress.ip_network(subnet, strict=False)
            
            # Get IP lists
            all_ips = list(network)                # All IPs including network/broadcast
            usable_ips = list(network.hosts())     # Usable IPs (excluding reserved)
            
            # Get network info
            network_address = str(network.network_address)
            broadcast_address = str(network.broadcast_address) if network.version == 4 else "N/A (IPv6)"
            netmask = str(network.netmask)
            
            # Get first and last usable IPs
            first_usable = str(usable_ips[0]) if usable_ips else "None"
            last_usable = str(usable_ips[-1]) if usable_ips else "None"
            
            # Network class (for IPv4)
            if network.version == 4:
                first_octet = int(network.network_address.packed[0])
                if first_octet < 128:
                    network_class = "Class A"
                elif first_octet < 192:
                    network_class = "Class B"
                elif first_octet < 224:
                    network_class = "Class C"
                elif first_octet < 240:
                    network_class = "Class D (Multicast)"
                else:
                    network_class = "Class E (Reserved)"
            else:
                network_class = "N/A (IPv6)"
            
            # Create the subnet display data (for Excel column)
            ip_list = [str(ip) for ip in all_ips]
            
            # Add separator and metadata at the bottom
            ip_list.append("")
            ip_list.append(f"Network ID: {network_address}")
            ip_list.append(f"Broadcast: {broadcast_address}")
            ip_list.append(f"Netmask: {netmask}")
            ip_list.append(f"Prefix: /{network.prefixlen}")
            ip_list.append("")
            ip_list.append(f"First Usable: {first_usable}")
            ip_list.append(f"Last Usable: {last_usable}")
            ip_list.append("")
            ip_list.append(f"Total IPs: {len(all_ips)}")
            ip_list.append(f"Usable IPs: {len(usable_ips)}")
            ip_list.append(f"Network Class: {network_class}")
            
            # Store for Excel
            subnet_data[subnet] = ip_list
            
            # Add to subnet summary
            subnet_summary.append({
                'Subnet': subnet,
                'IP Version': f"IPv{network.version}",
                'Network Class': network_class,
                'Prefix': f"/{network.prefixlen}",
                'Netmask': netmask,
                'Network ID': network_address,
                'Broadcast': broadcast_address,
                'First Usable': first_usable,
                'Last Usable': last_usable,
                'Total IPs': len(all_ips),
                'Usable IPs': len(usable_ips)
            })
            
            # Update statistics
            stats['valid_subnets'] += 1
            stats['total_ips'] += len(all_ips)
            stats['usable_ips'] += len(usable_ips)
            
        except Exception as e:
            print(f"  Warning: Invalid subnet '{subnet}': {str(e)}")
            subnet_data[subnet] = [f"INVALID SUBNET: {subnet}", f"Error: {str(e)}"]
            
            # Add to subnet summary with error
            subnet_summary.append({
                'Subnet': subnet,
                'IP Version': 'Invalid',
                'Network Class': 'Error',
                'Prefix': 'Error',
                'Netmask': 'Error',
                'Network ID': 'Error',
                'Broadcast': 'Error',
                'First Usable': 'Error',
                'Last Usable': 'Error',
                'Total IPs': 0,
                'Usable IPs': 0
            })
            
            stats['invalid_subnets'] += 1
    
    print_status(f"Analysis complete: {stats['valid_subnets']} valid subnets, {stats['invalid_subnets']} invalid")
    print_status(f"Total IPs: {stats['total_ips']} ({stats['usable_ips']} usable)")
    
    return subnet_data, subnet_summary, stats

def create_excel_report(subnet_data, subnet_summary, stats, output_file='subnet_ranges_detailed.xlsx'):
    """Create Excel report with enhanced formatting."""
    print_status(f"Creating Excel report: {output_file}")
    
    # Create DataFrames
    ip_df = pd.DataFrame(dict([(k, pd.Series(v)) for k, v in subnet_data.items()]))
    summary_df = pd.DataFrame(subnet_summary)
    
    # Create overall stats DataFrame
    overall_stats = pd.DataFrame({
        'Metric': [
            'Total Subnets',
            'Valid Subnets',
            'Invalid Subnets',
            'Total IP Addresses',
            'Total Usable IP Addresses',
            'Report Generated'
        ],
        'Value': [
            stats['total_subnets'],
            stats['valid_subnets'],
            stats['invalid_subnets'],
            stats['total_ips'],
            stats['usable_ips'],
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ]
    })
    
    # Write to Excel with 3 sheets and formatting
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Write the three sheets
            overall_stats.to_excel(writer, sheet_name='Overall Summary', index=False)
            summary_df.to_excel(writer, sheet_name='Subnet Summary', index=False)
            ip_df.to_excel(writer, sheet_name='IP Details', index=False)
            
            # Apply formatting
            workbook = writer.book
            
            # Format Overall Summary sheet
            ws = workbook['Overall Summary']
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # Format header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Format data rows
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center')
            
            # Set column widths
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 30
            
            # Format Subnet Summary sheet
            ws = workbook['Subnet Summary']
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Format data rows
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = thin_border
                    if "Error" in str(cell.value):
                        cell.font = Font(color="FF0000")  # Red font for errors
            
            # Set column widths
            for i in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(i)].width = 18
            
            # Format IP Details sheet
            ws = workbook['IP Details']
            for cell in ws[1]:  # Headers
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Set column widths
            for i in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(i)].width = 20
        
        print_status(f"Excel report created successfully: {output_file}")
        return True
    except Exception as e:
        print(f"Error creating Excel report: {str(e)}")
        return False

def main():
    """Main function to run the subnet analysis."""
    print("\n===== Subnet Analysis Tool =====\n")
    
    # Input and output files
    input_file = 'scope.txt'
    output_file = 'subnet_ranges_detailed.xlsx'
    
    # Read subnets
    subnets = read_subnets(input_file)
    
    # Analyze subnets
    subnet_data, subnet_summary, stats = analyze_subnets(subnets)
    
    # Create Excel report
    create_excel_report(subnet_data, subnet_summary, stats, output_file)
    
    print("\nAnalysis completed successfully!")
    print(f"Results saved to: {output_file}")

if __name__ == "__main__":
    main()